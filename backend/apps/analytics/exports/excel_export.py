from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, RadarChart, Reference


class AnalyticsExcelExport:

    @classmethod
    def export(cls, data):

        wb = Workbook()

        cls._sheet_summary(wb, data)

        cls._sheet_clustering(wb, data)

        cls._sheet_feature_importance(wb, data)

        cls._sheet_radar(wb, data)

        cls._sheet_topsis(wb, data)

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            )
        )

        response["Content-Disposition"] = (
            'attachment; filename="analytics_report.xlsx"'
        )

        wb.save(response)

        return response

    # =========================================================

    @staticmethod
    def _sheet_summary(wb, data):

        ws = wb.active

        ws.title = "Ringkasan"

        summary = data["summary"]

        ws.append(["Ringkasan Survey & Statistik Responden"])

        ws.append([])

        ws.append(["Jumlah Desa", summary["total_village"]])

        ws.append(["Jumlah Responden", summary["total_respondent"]])

        ws.append(["Jumlah Response", summary["total_response"]])

        ws.append(["Jumlah Cluster", summary["n_clusters"]])

        ws.append([

            "Silhouette Score",

            (
                round(summary["silhouette_score"], 4)
                if summary["silhouette_score"] is not None
                else "-"
            ),

        ])

        ws.append([])

        ws.append(["Kesimpulan Otomatis"])

        ws.append([data["narrative"]])

    @staticmethod
    def _sheet_clustering(wb, data):

        ws = wb.create_sheet("Hasil Clustering")

        ws.append([
            "Rank",
            "Desa",
            "Cluster",
            "Total Score",
        ])

        for row in data["village_table"]:

            ws.append([

                row["rank"],

                row["village"].name,

                row["cluster"].name if row["cluster"] else "-",

                float(row["total_score"]),

            ])

    @staticmethod
    def _sheet_feature_importance(wb, data):

        ws = wb.create_sheet("Feature Importance")

        ws.append(["Variabel", "Kontribusi (%)"])

        for item in data["variable_importance"]:

            ws.append([
                item["name"],
                item["percentage"],
            ])

        n_rows = len(data["variable_importance"])

        if n_rows == 0:
            return

        chart = BarChart()

        chart.title = "Indikator/Variabel Paling Dominan"

        chart.y_axis.title = "Kontribusi (%)"

        values = Reference(
            ws,
            min_col=2,
            min_row=1,
            max_row=n_rows + 1,
        )

        categories = Reference(
            ws,
            min_col=1,
            min_row=2,
            max_row=n_rows + 1,
        )

        chart.add_data(values, titles_from_data=True)

        chart.set_categories(categories)

        ws.add_chart(chart, "D2")

    @staticmethod
    def _sheet_radar(wb, data):

        ws = wb.create_sheet("Radar Chart")

        village = data["representative_village"]

        ws.append([

            "Radar Chart Desa Representatif: "
            f"{village.name if village else '-'}"

        ])

        ws.append(["Variabel", "Skor"])

        for axis in data["radar_data"]:

            ws.append([axis["name"], axis["score"]])

        n_rows = len(data["radar_data"])

        if n_rows == 0:
            return

        chart = RadarChart()

        chart.title = "Profil Indikator Dominan Desa"

        values = Reference(
            ws,
            min_col=2,
            min_row=2,
            max_row=n_rows + 2,
        )

        categories = Reference(
            ws,
            min_col=1,
            min_row=3,
            max_row=n_rows + 2,
        )

        chart.add_data(values, titles_from_data=True)

        chart.set_categories(categories)

        ws.add_chart(chart, "D2")

    @staticmethod
    def _sheet_topsis(wb, data):

        ws = wb.create_sheet("Ranking TOPSIS")

        ws.append([
            "Cluster",
            "Rank",
            "Desa",
            "Skor Preferensi",
        ])

        for group in data["cluster_ranking"]:

            cluster_name = (
                group["cluster"].name
                if group["cluster"]
                else "Belum Dikluster"
            )

            for item in group["ranking"]:

                ws.append([

                    cluster_name,

                    item["rank"],

                    item["village"].name,

                    (
                        round(item["score"], 4)
                        if item["score"] is not None
                        else "-"
                    ),

                ])
