import datetime
import re

from django.db import transaction
from openpyxl import load_workbook

from apps.master.models import Questionnaire, Village
from apps.respondent.models import Respondent
from apps.response.models import Response
from apps.survey.models import SurveyVillage


# Mapping ID Desa -> Nama Desa/Kelurahan (24 desa, Kota Batu),
# dipakai untuk mencocokkan kolom "Desa" di file Excel import
# (yang isinya angka) ke Village yang sudah ada di Master Data.
VILLAGE_ID_MAP = {
    1: "Oro-oro Ombo",
    2: "Ngaglik",
    3: "Pesanggrahan",
    4: "Songgokerto",
    5: "Sumberejo",
    6: "Temas",
    7: "Sisir",
    8: "Sidomulyo",
    9: "Bumiaji",
    10: "Punten",
    11: "Tulungrejo",
    12: "Sumbergondo",
    13: "Bulukerto",
    14: "Gunungsari",
    15: "Pandanrejo",
    16: "Giripurno",
    17: "Sumberbrantas",
    18: "Beji",
    19: "Torongrejo",
    20: "Mojorejo",
    21: "Pendem",
    22: "Junrejo",
    23: "Dadaprejo",
    24: "Tlekung",
}

# Kolom non-indikator yang dikenali di file Excel
META_COLUMNS = {
    "no",
    "id resp",
    "desa",
    "rt",
    "person",
}

DEFAULT_BIRTH_DATE = datetime.date(2000, 1, 1)

INDICATOR_PATTERN = re.compile(r"^[xy]\d+\.\d+$")


class BulkImportError(Exception):
    pass


class ResponseBulkImportService:
    """
    Import data historis (respondent + jawaban 88 indikator)
    dari file Excel, ke tabel Respondent & Response yang sudah
    ada -- TIDAK mengubah struktur/CRUD Respondent & Response.

    Kolom yang WAJIB ada di file: ID Resp, Desa.
    Kolom RT & Person opsional (disimpan di field `address`,
    karena model Respondent tidak punya kolom khusus untuk itu).
    Kolom lain (X1.1, X1.2, ..., Y6.8) dicocokkan ke Indicator
    lewat Questionnaire.indicator.code.

    Field wajib di model Respondent yang TIDAK ada di file
    (nik, gender, birth_date, address) akan diisi otomatis
    dengan nilai placeholder, karena datanya memang tidak
    tersedia di dataset historis.
    """

    @classmethod
    def import_excel(cls, file, survey):

        wb = load_workbook(
            filename=file,
            data_only=True,
        )

        ws = wb.active

        header_row, column_map = cls._find_header_columns(ws)

        questionnaire_map = cls._build_questionnaire_map(
            column_map,
        )

        village_cache = {}

        survey_village_cache = {}

        created_respondent = 0

        updated_respondent = 0

        created_response = 0

        skipped_rows = []

        for data_row in range(header_row + 1, ws.max_row + 1):

            row_dict = {

                name: ws.cell(row=data_row, column=col).value

                for name, col in column_map.items()

            }

            id_resp = row_dict.get("id resp")

            desa_id = row_dict.get("desa")

            if id_resp is None or desa_id is None:
                continue

            try:

                # 1 baris = 1 transaksi tersendiri, supaya kalau
                # ada 1 baris bermasalah, baris lain yang sudah
                # berhasil TIDAK ikut di-rollback.
                with transaction.atomic():

                    village = cls._resolve_village(
                        desa_id,
                        village_cache,
                    )

                    survey_village = cls._resolve_survey_village(
                        survey,
                        village,
                        survey_village_cache,
                    )

                    respondent, was_created = cls._resolve_respondent(
                        id_resp,
                        row_dict,
                        village,
                        survey_village,
                    )

                    answered = cls._import_answers(
                        respondent,
                        row_dict,
                        questionnaire_map,
                    )

            except BulkImportError as error:

                skipped_rows.append(
                    f"Baris {data_row}: {error}"
                )

                continue

            except Exception as error:

                skipped_rows.append(
                    f"Baris {data_row}: error tak terduga - {error}"
                )

                continue

            if was_created:
                created_respondent += 1
            else:
                updated_respondent += 1

            created_response += answered

        return {

            "created_respondent": created_respondent,

            "updated_respondent": updated_respondent,

            "created_response": created_response,

            "skipped_rows": skipped_rows,

        }

    # =========================================================

    @staticmethod
    def _find_header_columns(ws):
        """
        Cari sel "ID Resp" di seluruh sheet, lalu jalan ke KANAN
        kolom demi kolom SATU-PERSATU dari kolom itu untuk
        mengumpulkan kolom meta (ID Resp, Desa, RT, Person) dan
        indikator (X1.1 ... Y6.8), berhenti begitu ketemu kolom
        kosong atau header yang tidak dikenali.

        Ini SENGAJA tidak menyapu seluruh lebar sheet (yang bisa
        198 kolom berisi banyak tabel lain seperti Uji Validitas,
        Path Analysis, dst di sebelah kanan) supaya tidak salah
        cocok dengan teks "X1.1" dsb yang kebetulan muncul lagi
        di tabel lain pada baris yang sama.
        """

        id_resp_cell = None

        for row in ws.iter_rows(min_row=1, max_row=10):

            for cell in row:

                if (
                    cell.value is not None
                    and str(cell.value).strip().lower() == "id resp"
                ):

                    id_resp_cell = cell

                    break

            if id_resp_cell:
                break

        if id_resp_cell is None:

            raise BulkImportError(
                "Kolom \"ID Resp\" tidak ditemukan di file Excel."
            )

        header_row = id_resp_cell.row

        column_map = {}

        col = id_resp_cell.column

        blank_streak = 0

        while col <= ws.max_column:

            value = ws.cell(row=header_row, column=col).value

            text = (
                str(value).strip().lower()
                if value is not None
                else ""
            )

            if text == "":

                blank_streak += 1

                if blank_streak >= 1 and len(column_map) > 4:

                    break

            else:

                blank_streak = 0

                is_meta = text in META_COLUMNS

                is_indicator = bool(
                    INDICATOR_PATTERN.match(text)
                )

                if not is_meta and not is_indicator:
                    break

                column_map[text] = col

            col += 1

        if "id resp" not in column_map or "desa" not in column_map:

            raise BulkImportError(
                "Kolom \"ID Resp\" dan \"Desa\" wajib ada berdekatan "
                "di baris header."
            )

        return header_row, column_map

    @staticmethod
    def _build_questionnaire_map(column_map):

        indicator_codes = [

            name.upper()

            for name in column_map

            if name not in META_COLUMNS

        ]

        questionnaires = (
            Questionnaire.objects
            .filter(
                indicator__code__in=indicator_codes,
                is_active=True,
            )
            .select_related("indicator")
        )

        return {

            questionnaire.indicator.code.lower(): questionnaire

            for questionnaire in questionnaires

        }

    @classmethod
    def _resolve_village(cls, desa_id, cache):

        try:

            desa_id_int = int(desa_id)

        except (TypeError, ValueError):

            raise BulkImportError(
                f"Nilai kolom Desa \"{desa_id}\" bukan angka."
            )

        if desa_id_int in cache:
            return cache[desa_id_int]

        village_name = VILLAGE_ID_MAP.get(desa_id_int)

        if village_name is None:

            raise BulkImportError(
                f"ID Desa {desa_id_int} tidak dikenali."
            )

        village = (
            Village.objects
            .filter(name=village_name)
            .first()
        )

        if village is None:

            raise BulkImportError(
                f"Desa \"{village_name}\" belum ada di Master Data."
            )

        cache[desa_id_int] = village

        return village

    @staticmethod
    def _resolve_survey_village(survey, village, cache):

        key = village.id

        if key in cache:
            return cache[key]

        survey_village, _ = SurveyVillage.objects.get_or_create(
            survey=survey,
            village=village,
        )

        cache[key] = survey_village

        return survey_village

    @staticmethod
    def _resolve_respondent(
        id_resp,
        row_dict,
        village,
        survey_village,
    ):

        nik = f"IMPORT-{int(id_resp)}"

        rt = row_dict.get("rt", "")

        person = row_dict.get("person", "")

        address = (
            f"{village.name}, RT {rt} - Jabatan: {person}"
        )

        respondent, was_created = Respondent.objects.update_or_create(
            nik=nik,
            defaults={
                "survey_village": survey_village,
                "name": f"Responden Import #{int(id_resp)}",
                "gender": "M",
                "birth_date": DEFAULT_BIRTH_DATE,
                "address": address,
            },
        )

        return respondent, was_created

    @staticmethod
    def _import_answers(respondent, row_dict, questionnaire_map):

        count = 0

        for code_lower, questionnaire in questionnaire_map.items():

            value = row_dict.get(code_lower)

            if value is None or value == "":
                continue

            answer_type = questionnaire.answer_type

            defaults = {
                "answer_boolean": None,
                "answer_integer": None,
                "answer_decimal": None,
                "answer_text": "",
                "score": 0,
            }

            try:

                if answer_type == "boolean":

                    defaults["answer_boolean"] = bool(int(value))

                    defaults["score"] = (
                        5 if int(value) == 1 else 1
                    )

                elif answer_type in ("likert", "integer"):

                    defaults["answer_integer"] = int(value)

                    defaults["score"] = int(value)

                elif answer_type == "decimal":

                    defaults["answer_decimal"] = float(value)

                    defaults["score"] = float(value)

                else:

                    defaults["answer_text"] = str(value)

                    defaults["score"] = 0

            except (TypeError, ValueError):

                # Nilai di sel ini tidak sesuai tipe yang
                # diharapkan (misal ternyata teks, bukan angka) --
                # lewati jawaban ini saja, jangan gagalkan
                # seluruh baris/import.
                continue

            Response.objects.update_or_create(
                respondent=respondent,
                questionnaire=questionnaire,
                defaults=defaults,
            )

            count += 1

        return count
